import os
from datetime import datetime
from io import BytesIO

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from core.normalization import normalize_column_name


def load_input_file(uploaded_file) -> pd.DataFrame:
    name = uploaded_file.name.lower()

    if name.endswith(".csv"):
        df = pd.read_csv(uploaded_file)
    elif name.endswith((".xlsx", ".xls")):
        df = pd.read_excel(uploaded_file)
    else:
        raise ValueError("File type not supported. Use CSV or Excel (.csv, .xlsx, .xls).")

    df.columns = [normalize_column_name(col) for col in df.columns]
    return df


def to_excel_autofit(df: pd.DataFrame) -> bytes:
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
        worksheet = writer.sheets["Sheet1"]

        for col_idx, col in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            max_length = len(str(col))
            for cell in worksheet[col_letter]:
                cell_value = str(cell.value) if cell.value is not None else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            worksheet.column_dimensions[col_letter].width = max_length + 2

    output.seek(0)
    return output.getvalue()


def _safe_sheet_name(name: str, used_names: set[str]) -> str:
    invalid_chars = ["\\", "/", "*", "[", "]", ":", "?"]
    safe_name = str(name or "Sheet").strip()

    for char in invalid_chars:
        safe_name = safe_name.replace(char, "-")

    safe_name = safe_name[:31] or "Sheet"
    candidate = safe_name
    suffix = 1

    while candidate in used_names:
        suffix_text = f" {suffix}"
        candidate = f"{safe_name[:31 - len(suffix_text)]}{suffix_text}"
        suffix += 1

    used_names.add(candidate)
    return candidate


def to_excel_workbook_autofit(sheets: dict[str, pd.DataFrame]) -> bytes:
    output = BytesIO()
    used_sheet_names = set()
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    total_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    currency_format = '£#,##0.00'

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for requested_sheet_name, df in sheets.items():
            sheet_name = _safe_sheet_name(requested_sheet_name, used_sheet_names)
            export_df = df.copy() if df is not None else pd.DataFrame()
            export_df.to_excel(writer, index=False, sheet_name=sheet_name)
            worksheet = writer.sheets[sheet_name]
            worksheet.freeze_panes = "A2"

            if export_df.shape[1] > 0:
                last_col = get_column_letter(export_df.shape[1])
                worksheet.auto_filter.ref = f"A1:{last_col}{max(1, export_df.shape[0] + 1)}"

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font

            for col_idx, col in enumerate(export_df.columns, start=1):
                col_letter = get_column_letter(col_idx)
                max_length = len(str(col))
                is_currency_col = "price" in str(col).lower() or "total" in str(col).lower()

                for cell in worksheet[col_letter]:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)

                    if cell.row > 1 and is_currency_col:
                        cell.number_format = currency_format

                worksheet.column_dimensions[col_letter].width = min(max_length + 2, 60)

            if "Product Item" in export_df.columns:
                product_item_col = list(export_df.columns).index("Product Item") + 1
                for row_idx in range(2, export_df.shape[0] + 2):
                    if worksheet.cell(row_idx, product_item_col).value == "TOTAL":
                        for cell in worksheet[row_idx]:
                            cell.fill = total_fill
                            cell.font = Font(bold=True)

    output.seek(0)
    return output.getvalue()


def build_output_filenames() -> tuple[str, str]:
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    return (
        f"dpl_output_{stamp}.csv",
        f"dpl_output_{stamp}.xlsx",
    )


def build_tracking_output_filename(original_name: str) -> str:
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    base, ext = os.path.splitext(original_name)
    return f"{base}_with_tracking_{stamp}{ext.lower()}"


def dataframe_to_download_bytes(df: pd.DataFrame, original_name: str) -> tuple[bytes, str, str]:
    lower = original_name.lower()

    if lower.endswith(".csv"):
        return (
            df.to_csv(index=False).encode("utf-8"),
            build_tracking_output_filename(original_name),
            "text/csv",
        )

    if lower.endswith((".xlsx", ".xls")):
        return (
            to_excel_autofit(df),
            build_tracking_output_filename(original_name),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    raise ValueError("Unsupported output file type.")


def get_file_type(filename: str) -> str:
    lower = filename.lower()
    if lower.endswith(".csv"):
        return "csv"
    if lower.endswith(".xlsx"):
        return "xlsx"
    if lower.endswith(".xls"):
        return "xls"
    if lower.endswith(".pdf"):
        return "pdf"
    return "unknown"
