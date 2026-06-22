from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from core.file_io import to_excel_workbook_autofit


def test_to_excel_workbook_autofit_writes_multiple_named_sheets():
    workbook_bytes = to_excel_workbook_autofit(
        {
            "Summary": pd.DataFrame([{"Category": "Total orders", "Count": 2}]),
            "Billing Details": pd.DataFrame(
                [
                    {
                        "Product Item": "Adult T-Shirt",
                        "Item Price": 5.5,
                        "Shipping Price": 2.8,
                        "Line Total": "=B2+C2",
                    },
                    {
                        "Product Item": "TOTAL",
                        "Item Price": "",
                        "Shipping Price": "",
                        "Line Total": "=SUM(D2:D2)",
                    },
                ]
            ),
            "This sheet name is deliberately longer than thirty one characters": pd.DataFrame(
                [{"Item": "Mug", "Count": 2}]
            ),
        }
    )

    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)

    assert workbook.sheetnames[0] == "Summary"
    assert workbook.sheetnames[2] == "This sheet name is deliberately"
    assert workbook["Summary"]["A1"].value == "Category"
    assert workbook["Summary"].freeze_panes == "A2"
    assert workbook["Billing Details"]["B2"].number_format == "£#,##0.00"
    assert workbook["Billing Details"]["A3"].font.bold is True
